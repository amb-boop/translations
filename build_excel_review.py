import csv
import json
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.worksheet.datavalidation import DataValidation


ROOT = Path(__file__).resolve().parent
OUTPUT = ROOT / "Etam_Localization_Review.xlsx"

COUNTRY_FILES = {
    "Spain": "es_translation_queue.csv",
    "Poland": "pl_translation_queue.csv",
    "Czech Republic": "cz_translation_queue.csv",
    "United Kingdom": "uk_translation_queue.csv",
}

REVIEW_COLUMNS = [
    "key",
    "run_date",
    "locale",
    "country",
    "reference_mc",
    "model_id",
    "color_id",
    "change_type",
    "confidence",
    "product_url",
    "image_url",
    "category",
    "universe",
    "current_local_title",
    "current_local_long_description",
    "proposed_title",
    "proposed_long_description",
    "source_logic",
    "proposal_source",
    "fr_title_changed",
    "fr_long_description_changed",
    "previous_fr_title",
    "current_fr_title",
    "previous_fr_long_description",
    "current_fr_long_description",
    "status",
    "title_validation",
    "description_validation",
    "reviewer_comment",
    "approved_title",
    "approved_long_description",
    "last_reviewed_at",
]

PRESERVED_COLUMNS = [
    "status",
    "title_validation",
    "description_validation",
    "reviewer_comment",
    "approved_title",
    "approved_long_description",
    "last_reviewed_at",
]


def read_csv_rows(filename):
    path = ROOT / filename
    if not path.exists():
        return []
    with path.open("r", encoding="utf-8", newline="") as handle:
        return list(csv.DictReader(handle, delimiter=";"))


def row_key(row):
    return f"{row.get('locale', '')}|{row.get('reference_mc', '')}|{row.get('change_type', '')}"


def read_preserved_state():
    if not OUTPUT.exists():
        return {}
    workbook = load_workbook(OUTPUT, read_only=True, data_only=False)
    preserved = {}
    for sheet_name in COUNTRY_FILES:
        if sheet_name not in workbook.sheetnames:
            continue
        sheet = workbook[sheet_name]
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            continue
        header_row = next((row for row in rows if "key" in [str(value or "") for value in row]), None)
        if not header_row:
            continue
        header_index = rows.index(header_row)
        headers = [str(value or "") for value in header_row]
        if "key" not in headers:
            continue
        key_index = headers.index("key")
        indexes = {column: headers.index(column) for column in PRESERVED_COLUMNS if column in headers}
        for row in rows[header_index + 1:]:
            key = row[key_index]
            if not key:
                continue
            preserved[str(key)] = {
                column: row[indexes[column]] if indexes[column] < len(row) else ""
                for column in indexes
            }
    workbook.close()
    return preserved


def style_header(sheet, column_count, row_number=1):
    fill = PatternFill("solid", fgColor="F5EDEB")
    for cell in sheet[row_number][:column_count]:
        cell.fill = fill
        cell.font = Font(bold=True, color="25282D")
        cell.alignment = Alignment(vertical="top", wrap_text=True)


def set_widths(sheet):
    widths = {
        "A": 28,
        "B": 12,
        "C": 10,
        "D": 16,
        "E": 13,
        "H": 28,
        "I": 10,
        "J": 36,
        "K": 36,
        "N": 28,
        "O": 42,
        "P": 28,
        "Q": 46,
        "R": 42,
        "V": 28,
        "W": 28,
        "X": 42,
        "Y": 42,
        "Z": 16,
        "AA": 28,
        "AB": 28,
        "AC": 46,
        "AD": 18,
    }
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width


def add_review_validation(sheet, header_row):
    for column_name in ["status", "title_validation", "description_validation"]:
        column_index = REVIEW_COLUMNS.index(column_name) + 1
        validation = DataValidation(
            type="list",
            formula1='"Draft,Approved,Rejected,Needs copy review"',
            allow_blank=False,
        )
        sheet.add_data_validation(validation)
        validation.add(f"{sheet.cell(header_row + 1, column_index).coordinate}:{sheet.cell(max(sheet.max_row, header_row + 1), column_index).coordinate}")


def write_review_sheet(workbook, name, rows, preserved, metrics):
    sheet = workbook.create_sheet(name)
    sheet["A1"] = f"{name} recap"
    sheet["A1"].font = Font(bold=True, size=14, color="25282D")
    sheet["A2"] = f"{metrics.get('mc_to_review', len(rows))} MC references to review"
    sheet["A3"] = f"{metrics.get('title_translation_issues', 0)} title translation issues"
    sheet["A4"] = f"{metrics.get('missing_long_descriptions', 0)} missing long descriptions"
    sheet["A5"] = f"{metrics.get('same_model_colorway_reusable_descriptions', 0)} can reuse same-model/colorway descriptions"
    sheet["D2"] = "Validation: set status to Approved for the full row, or approve title/description separately."
    sheet["D2"].alignment = Alignment(wrap_text=True, vertical="top")
    sheet["D2"].font = Font(color="6E6663")
    header_row = 7
    while sheet.max_row < header_row - 1:
        sheet.append([])
    sheet.append(REVIEW_COLUMNS)
    for source in rows:
        key = row_key(source)
        saved = preserved.get(key, {})
        source = dict(source)
        source["key"] = key
        source.setdefault("status", "Draft")
        source.setdefault("title_validation", "Draft")
        source.setdefault("description_validation", "Draft")
        values = []
        for column in REVIEW_COLUMNS:
            if column in PRESERVED_COLUMNS and saved.get(column):
                values.append(saved[column])
            else:
                values.append(source.get(column, ""))
        sheet.append(values)
    sheet.freeze_panes = f"A{header_row + 1}"
    sheet.auto_filter.ref = f"A{header_row}:{sheet.cell(max(sheet.max_row, header_row), len(REVIEW_COLUMNS)).coordinate}"
    style_header(sheet, len(REVIEW_COLUMNS), header_row)
    set_widths(sheet)
    add_review_validation(sheet, header_row)
    for row in sheet.iter_rows(min_row=header_row + 1):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def write_summary(workbook, summary, country_rows):
    sheet = workbook.active
    sheet.title = "Summary"
    sheet.append(["Metric", "Value"])
    sheet.append(["Run date", summary.get("run_date", "")])
    sheet.append(["Mode", summary.get("mode", "")])
    sheet.append(["France MC references", summary.get("france", {}).get("total_mc_references", 0)])
    sheet.append(["Queue items", sum(len(rows) for rows in country_rows.values())])
    sheet.append(["Status", summary.get("first_run_message", "")])
    sheet.append([])
    sheet.append(["Country", "MC to review", "Title issues", "Missing long descriptions", "Same model reusable"])
    for locale, country in summary.get("countries", {}).items():
        metrics = country.get("metrics", {})
        sheet.append([
            country.get("name", locale),
            metrics.get("mc_to_review", 0),
            metrics.get("title_translation_issues", 0),
            metrics.get("missing_long_descriptions", 0),
            metrics.get("same_model_colorway_reusable_descriptions", 0),
        ])
    style_header(sheet, 5)
    sheet.column_dimensions["A"].width = 34
    sheet.column_dimensions["B"].width = 18
    sheet.column_dimensions["C"].width = 18
    sheet.column_dimensions["D"].width = 28
    sheet.column_dimensions["E"].width = 24


def write_daily_changes(workbook, summary):
    sheet = workbook.create_sheet("Daily changes")
    countries = summary.get("countries", {})
    sheet.append([
        "run_date",
        "new_mc_references_today",
        "french_titles_changed",
        "french_long_descriptions_changed",
        "country_content_gaps_detected",
        "persistent_issues",
        "resolved_since_last_run",
        "products_removed_from_feed",
    ])
    sheet.append([
        summary.get("run_date", ""),
        summary.get("france", {}).get("new_mc_references_today", 0),
        summary.get("france", {}).get("french_titles_changed", 0),
        summary.get("france", {}).get("french_long_descriptions_changed", 0),
        sum(country.get("metrics", {}).get("country_content_gaps_detected", 0) for country in countries.values()),
        sum(country.get("metrics", {}).get("persistent_issues", 0) for country in countries.values()),
        sum(country.get("metrics", {}).get("resolved_since_last_run", 0) for country in countries.values()),
        summary.get("france", {}).get("products_removed_from_feed", 0),
    ])
    style_header(sheet, 8)
    for index in range(1, 9):
        sheet.column_dimensions[chr(64 + index)].width = 24


def write_history(workbook, previous_history, summary, country_rows):
    sheet = workbook.create_sheet("History")
    headers = ["timestamp", "run_date", "mode", "queue_items", "message"]
    sheet.append(headers)
    for row in previous_history:
        sheet.append(row)
    sheet.append([
        summary.get("run_date", ""),
        summary.get("run_date", ""),
        summary.get("mode", ""),
        sum(len(rows) for rows in country_rows.values()),
        summary.get("first_run_message", ""),
    ])
    style_header(sheet, len(headers))
    sheet.column_dimensions["A"].width = 16
    sheet.column_dimensions["B"].width = 14
    sheet.column_dimensions["C"].width = 18
    sheet.column_dimensions["D"].width = 14
    sheet.column_dimensions["E"].width = 90


def read_previous_history():
    if not OUTPUT.exists():
        return []
    workbook = load_workbook(OUTPUT, read_only=True, data_only=True)
    if "History" not in workbook.sheetnames:
        workbook.close()
        return []
    rows = list(workbook["History"].iter_rows(values_only=True))
    workbook.close()
    return [list(row) for row in rows[1:] if any(row)]


def main():
    summary_path = ROOT / "summary.json"
    summary = json.loads(summary_path.read_text(encoding="utf-8")) if summary_path.exists() else {}
    preserved = read_preserved_state()
    previous_history = read_previous_history()
    country_rows = {name: read_csv_rows(filename) for name, filename in COUNTRY_FILES.items()}

    workbook = Workbook()
    write_summary(workbook, summary, country_rows)
    write_daily_changes(workbook, summary)
    for name, rows in country_rows.items():
        metrics = next(
            (country.get("metrics", {}) for country in summary.get("countries", {}).values() if country.get("name") == name),
            {},
        )
        write_review_sheet(workbook, name, rows, preserved, metrics)
    write_history(workbook, previous_history, summary, country_rows)

    workbook.save(OUTPUT)
    print(OUTPUT)


if __name__ == "__main__":
    main()
