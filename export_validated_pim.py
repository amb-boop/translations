import csv
import re
from datetime import datetime
from pathlib import Path
import unicodedata

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parent
WORKBOOK_CANDIDATES = [
    ROOT / "Etam_Localization_Review.xlsx",
    ROOT / "Etam_PIM_Localization_Validation_Tool.xlsx",
]
EXPORT_DIR = ROOT / "exports" / datetime.now().strftime("%Y-%m-%d")

COUNTRY_SHEETS = {
    "Spain": "es_ES",
    "Poland": "pl_PL",
    "Czech Republic": "cz_CZ",
    "United Kingdom": "en_UK",
}

PIM_LOCALES = {
    "es_ES": "es_ES",
    "pl_PL": "pl_PL",
    "cz_CZ": "cz_CZ",
    "en_UK": "en_GB",
}

PIM_COLUMNS = [
    "code MC",
    "locale",
    "displayName fr",
    "shortDescription fr",
    "longDescription fr",
    "coreGlobalModelDescription fr",
    "productColor FR",
    "whyWeLoveIt fr",
    "ourStyleAdvice fr",
    "productDetails fr",
    "weCareDescription fr",
    "ourProduct fr",
    "confectionAndTransparency fr",
    "respectAnimalWelfare fr",
]


def normalize(value):
    if value is None:
        return ""
    text = str(value).replace("\ufeff", "").replace("\x00", "")
    text = text.replace("\r\n", "\n").replace("\r", "\n").strip()
    return unicodedata.normalize("NFC", text)


def is_approved(value):
    return normalize(value).lower() == "approved"


def pim_rich_text(value):
    text = normalize(value)
    text = text.replace("\r\n", "\n").replace("\r", "\n")
    text = re.sub(r"\s*•\s*", "\n• ", text)
    text = re.sub(r"\n{3,}", "\n\n", text).strip()
    return text.replace("\n", "<br>")


def read_sheet_rows(sheet):
    values = list(sheet.iter_rows(values_only=True))
    if not values:
        return []
    header_index = None
    headers = []
    for index, row in enumerate(values):
        candidate = [normalize(value) for value in row]
        if "key" in candidate and "reference_mc" in candidate:
            header_index = index
            headers = candidate
            break
    if header_index is None:
        return []
    rows = []
    for raw in values[header_index + 1:]:
        if not any(raw):
            continue
        rows.append({headers[index]: raw[index] if index < len(raw) else "" for index in range(len(headers))})
    return rows


def pim_row(row, locale):
    status_approved = is_approved(row.get("status"))
    title_approved = status_approved or is_approved(row.get("title_validation"))
    description_approved = status_approved or is_approved(row.get("description_validation"))

    if not title_approved and not description_approved:
        return None

    approved_title = normalize(row.get("approved_title")) or normalize(row.get("proposed_title"))
    approved_description = normalize(row.get("approved_long_description")) or normalize(row.get("proposed_long_description"))

    output = {column: "" for column in PIM_COLUMNS}
    output["code MC"] = normalize(row.get("reference_mc"))
    output["locale"] = PIM_LOCALES.get(locale, locale)

    if title_approved:
        output["displayName fr"] = approved_title
        output["shortDescription fr"] = approved_title
    if description_approved:
        output["longDescription fr"] = pim_rich_text(approved_description)

    return output


def write_csv(path, rows):
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=PIM_COLUMNS, delimiter=";", quoting=csv.QUOTE_MINIMAL)
        writer.writeheader()
        writer.writerows(rows)


def main():
    workbook_path = next((path for path in WORKBOOK_CANDIDATES if path.exists()), None)
    if workbook_path is None:
        expected = ", ".join(str(path) for path in WORKBOOK_CANDIDATES)
        raise SystemExit(f"Workbook not found. Expected one of: {expected}")

    workbook = load_workbook(workbook_path, read_only=True, data_only=False)
    all_rows = []
    export_counts = {}

    for sheet_name, locale in COUNTRY_SHEETS.items():
        if sheet_name not in workbook.sheetnames:
            export_counts[locale] = 0
            export_locale = PIM_LOCALES.get(locale, locale)
            write_csv(EXPORT_DIR / f"pim_import_{export_locale}_{datetime.now().strftime('%Y%m%d')}.csv", [])
            continue
        rows = []
        for row in read_sheet_rows(workbook[sheet_name]):
            output = pim_row(row, locale)
            if output:
                rows.append(output)
                all_rows.append(output)
        export_counts[locale] = len(rows)
        export_locale = PIM_LOCALES.get(locale, locale)
        write_csv(EXPORT_DIR / f"pim_import_{export_locale}_{datetime.now().strftime('%Y%m%d')}.csv", rows)

    write_csv(EXPORT_DIR / f"pim_import_all_validated_{datetime.now().strftime('%Y%m%d')}.csv", all_rows)
    workbook.close()

    print(f"PIM export folder: {EXPORT_DIR}")
    for locale, count in export_counts.items():
        print(f"{locale}: {count} validated rows")
    print(f"all: {len(all_rows)} validated rows")


if __name__ == "__main__":
    main()
